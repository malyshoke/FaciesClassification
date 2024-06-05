import sys
import os
from datetime import datetime
from io import BytesIO
import requests
from sqlalchemy import create_engine, Column, Integer, String, Float, ForeignKey, DateTime
import openpyxl
import pandas as pd
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QListWidgetItem
from PySide6.QtCore import QUrl, Qt
from openpyxl.utils import get_column_letter

import constants
from LithologyModel import LithologyModel
from ui_mainwindow import Ui_MainWindow
from cegal.welltools.wells import Well
from cegal.welltools.plotting import CegalWellPlotter as cwp
from plotly.offline import plot
from constants import Constants
from PySide6.QtWidgets import QMessageBox
import re
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl.styles import Font, Alignment, PatternFill


class LithologyDescription:
    def __init__(self, lithology_dict):
        self.dictionary = lithology_dict


def add_legend_to_figure(fig, lithology_description, well_name):
    # Получаем количество колонок из меток осей x
    x_axes = [ax for ax in fig.layout if re.match(r'xaxis\d*', ax)]
    num_cols = len(x_axes)
    lithology_values = list(lithology_description.dictionary.keys())
    lithology_names = [
        "Песчаник", "Песчан/Сланец", "Сланец", "Мергель", "Доломит", "Известняк",
        "Мел", "Галит", "Ангидрит", "Туф", "Уголь"
    ]
    lithology_colors = [lithology_description.dictionary[val][1] for val in lithology_values]

    # Создаем новую фигуру с дополнительной колонкой для легенды
    fig_with_legend = make_subplots(
        rows=1,
        cols=num_cols + 1,  # Увеличиваем количество колонок на 1
        shared_yaxes=True,
        subplot_titles=[*[fig.layout.annotations[i].text for i in range(len(fig.layout.annotations))], "Facies Colors"]
    )

    # Копируем существующие трейсы в новую фигуру (без легенды)
    for trace in fig.data:
        col = int(re.search(r'\d+', trace.xaxis).group()) if re.search(r'\d+', trace.xaxis) else 1
        trace.showlegend = False  # Убираем легенду
        fig_with_legend.add_trace(trace, row=1, col=col)

    # Создаем Heatmap для легенды литологии
    fig_with_legend.add_trace(
        go.Heatmap(
            z=np.array(lithology_values).reshape(len(lithology_values), 1),
            # Преобразуем значения литологии в 2D массив
            x=['Facies Colors'],  # Используем подпись для нового столбца
            y=lithology_names,
            colorscale=list(zip(np.linspace(0, 1, len(lithology_colors)), lithology_colors)),
            showscale=False,  # Отключаем цветовую шкалу
            hoverinfo='none'
        ),
        row=1,
        col=num_cols + 1  # Добавляем в новую последнюю колонку
    )

    # Проверка и обновление ширины
    current_width = fig.layout.width if fig.layout.width is not None else 700  # Задаем значение по умолчанию, если None

    # Обновляем макет
    fig_with_legend.update_layout(
        height=fig.layout.height,
        width=current_width + 250,  # Увеличиваем ширину для новой колонки
        title='',  # Убираем заголовок
        yaxis=dict(autorange='reversed')  # Инвертируем ось Y
    )

    # Поворачиваем названия колонок на 45 градусов
    for annotation in fig_with_legend.layout.annotations:
        annotation['textangle'] = 45

    # Добавляем новые аннотации справа
    annotations = []
    for i, name in enumerate(lithology_names):
        annotations.append(
            dict(
                x=1,  # Располагаем немного за пределами графика
                y=(i + 0.5) / len(lithology_names),
                xref='paper',
                yref='paper',
                text=name,
                showarrow=False,
                font=dict(size=12),
                xanchor='left',
                yanchor='middle'
            )
        )

    # Объединяем существующие и новые аннотации
    existing_annotations = list(fig_with_legend.layout.annotations) + annotations

    fig_with_legend.update_layout(annotations=existing_annotations)

    fig_with_legend.update_layout(
        width=current_width + 250,  # Увеличиваем ширину для новой колонки
        height=720,
        title_text=f"Скважина {well_name}",
        title_x=0.5
    )

    fig_with_legend.update_layout(showlegend=False)

    return fig_with_legend



def save_to_excel(dataframe, well_name, file_path, title, exclude_columns):
    # Исключение нежелательных колонок
    dataframe = dataframe.drop(columns=exclude_columns, errors='ignore')
    if 'DEPTH_MD' in dataframe.columns:
        columns = ['DEPTH_MD'] + [col for col in dataframe.columns if col != 'DEPTH_MD']
        dataframe = dataframe[columns]

    dataframe = dataframe.iloc[:, :-2]

    # Создание новой книги Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Объединение ячеек для заголовка
    ws.merge_cells('A1:Q1')
    title_cell = ws.cell(row=1, column=1, value=f"Скважина {well_name} - {title}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Настройка стилей для заголовков колонок
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Заполнение заголовков колонок
    headers = list(dataframe.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_num)].width = 12

    # Заполнение данных из DataFrame
    for row_num, row_data in enumerate(dataframe.itertuples(index=False), 3):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Сохранение файла
    wb.save(file_path)
    print(f"Отчет сохранен в {file_path}")



class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.file_path = None
        self.columns = []
        self.test_well = None
        self.lithology_model = LithologyModel('ovr_classifier.pkl', 'scaler.pkl')


        lithology_description_dict = {
            Constants.LITHOLOGY_NUMBERS[v]: (k, Constants.LITHOLOGY_COLORS[Constants.LITHOLOGY_NUMBERS[v]])
            for k, v in Constants.LITHOLOGY_KEYS.items()
        }
        self.lithology_description = LithologyDescription(lithology_description_dict)
        self.ui.saveButtonFile.clicked.connect(self.save_predictions_to_excel)
        self.ui.pushButtonLoad.clicked.connect(self.load_las_file)
        self.ui.pushButtonFilePlot.clicked.connect(self.plot_full_graph)
        self.ui.pushButtonFilePredict.clicked.connect(self.get_lithology_predictions2)
        self.ui.saveToDbButtonFile.clicked.connect(self.save_to_db_file)

        self.ui.saveButtonDb.clicked.connect(self.save_predictions_to_excel)
        self.ui.wellComboBox.currentIndexChanged.connect(self.on_well_selected)
        self.ui.pushButtonDbPlot.clicked.connect(self.plot_full_graph1)
        self.ui.pushButtonDbPredict.clicked.connect(self.get_lithology_predictions)
        self.ui.plotWellButton.clicked.connect(self.get_well_data_and_plot)
        self.ui.tabWidget.currentChanged.connect(self.on_tab_changed)
        self.load_well_names()

    def load_well_data(self, well_name):
        try:
            url = f'http://127.0.0.1:8081/well_data?well_name={well_name}'
            response = requests.get(url)

            if response.status_code == 200:
                data = response.json()
                self.test_well = pd.DataFrame(data)
                self.file_path = f"{well_name}.las"  # Имя файла для последующего использования

                # Выбираем только те колонки, которые не входят в exclude_columns
                self.columns = [col for col in self.test_well.columns if col not in Constants.exclude_columns]
                self.ui.columnsListWidgetDb.clear()
                for column in self.columns:
                    item = QListWidgetItem(column)
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                    item.setCheckState(Qt.Unchecked)
                    self.ui.columnsListWidgetDb.addItem(item)
            else:
                print(f"Ошибка при получении данных: {response.status_code}")

        except Exception as e:
            print(f"Ошибка при выполнении запроса: {str(e)}")

    def load_well_names(self):
        try:
            url = 'http://127.0.0.1:8081/wells_names'
            response = requests.get(url)

            if response.status_code == 200:
                wells = response.json()
                self.ui.wellComboBox.clear()  # Очищаем выпадающий список перед добавлением новых элементов
                self.ui.wellComboBox.addItems(wells)
            else:
                print(f"Ошибка при получении данных: {response.status_code}")

        except Exception as e:
            print(f"Ошибка при выполнении запроса: {str(e)}")

    def on_well_selected(self, index):
        well_name = self.ui.wellComboBox.currentText()
        self.load_well_data(well_name)

    def save_to_db_file(self):
        if self.file_path:
            try:
                url = 'http://127.0.0.1:8081/save_to_db'
                files = {'file': open(self.file_path, 'rb')}
                response = requests.post(url, files=files)

                if response.status_code == 200:
                    QMessageBox.information(self, "Успех", "Файл успешно сохранен в базе данных.")
                else:
                    QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении файла: {response.json().get('error')}")

            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при выполнении запроса: {str(e)}")
        else:
            QMessageBox.warning(self, "Файл не выбран",
                                "Пожалуйста, выберите файл LAS перед сохранением в базу данных.")
    def plot_full_graph1(self, well_name):
        try:
            url = f'http://127.0.0.1:8081/well_data?well_name={well_name}'
            response = requests.get(url)

            if response.status_code == 200:
                data = response.json()
                self.test_well = pd.DataFrame(data)
                self.file_path = f"{well_name}.las"  # Имя файла для последующего использования

                # Выбираем только те колонки, которые не входят в exclude_columns
                self.columns = [col for col in self.test_well.columns if col not in Constants.exclude_columns]
                self.ui.columnsListWidgetDb.clear()
                for column in self.columns:
                    item = QListWidgetItem(column)
                    item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                    item.setCheckState(Qt.Unchecked)
                    self.ui.columnsListWidgetDb.addItem(item)

                self.plot_full_graph()

            else:
                print(f"Ошибка при получении данных: {response.status_code}")

        except Exception as e:
            print(f"Ошибка при выполнении запроса: {str(e)}")


    def on_tab_changed(self, index):
        if self.ui.tabWidget.tabText(index) == "Загрузка из БД":
            self.load_well_names()

    def load_las_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Выбрать LAS файл", "", "LAS Files (*.las)")
        if file_path:
            self.file_path = file_path
            print(f"Файл выбран: {file_path}")

            force_well = Well(filename=self.file_path, path=None)
            self.test_well = pd.DataFrame(force_well.df())

            # Выбираем только те колонки, которые не входят в exclude_columns
            self.columns = [col for col in self.test_well.columns if col not in Constants.exclude_columns]
            self.ui.columnsListWidgetFile.clear()
            for column in self.columns:
                item = QListWidgetItem(column)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Unchecked)
                self.ui.columnsListWidgetFile.addItem(item)

    def save_predictions_to_excel(self):
        if self.test_well is not None and self.predictions is not None:
            # Извлечение названия скважины из пути к файлу
            well_name = os.path.basename(self.file_path).replace(".las", "")
            default_file_name = f"{well_name}.xlsx"
            file_name, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", default_file_name,
                                                       "Excel Files (*.xlsx)")
            if file_name:
                if not file_name.endswith(".xlsx"):
                    file_name += ".xlsx"
                test_well = self.test_well
                print(test_well)
                save_to_excel(test_well, well_name, file_name, "GeoLogML Report",
                              exclude_columns=Constants.exclude_columns[1:])
                print(f"Отчет сохранен в {file_name}")
        else:
            QMessageBox.warning(self, "Нет данных",
                                "Загрузите данные и получите предсказания перед сохранением в Excel.")

    def get_selected_columns(self):
        selected_columns = []
        for index in range(self.ui.columnsListWidgetFile.count()):
            item = self.ui.columnsListWidgetFile.item(index)
            if item.checkState() == Qt.Checked:
                selected_columns.append(item.text())
        return selected_columns

    def get_well_data_and_plot(self):
        try:
            url = 'http://127.0.0.1:8081/well_loc'
            response = requests.get(url)

            if response.status_code == 200:
                data = response.json()
                df = pd.DataFrame(data)
                self.plot_well_data(df)
            else:
                print(f"Ошибка при получении данных: {response.status_code}")

        except Exception as e:
            print(f"Ошибка при выполнении запроса: {str(e)}")

    def plot_well_data(self, data):
        # Преобразование данных в DataFrame
        df = pd.DataFrame(data)

        # Генерация уникальных цветов для каждой скважины
        unique_wells = df['WELL'].unique()
        color_map = {well: f'rgb({i * 50 % 256}, {(i * 50 + 100) % 256}, {(i * 50 + 200) % 256})' for i, well in
                     enumerate(unique_wells)}
        colors = df['WELL'].map(color_map)

        fig = go.Figure()

        fig.add_trace(go.Scatter(
            x=df['X_LOC'],
            y=df['Y_LOC'],
            mode='markers+text',
            marker=dict(size=df['data_points'], sizemode='area', sizeref=2. * max(df['data_points']) / (40. ** 2),
                        sizemin=4, color=colors),
            text=df['WELL'],
            textposition='top center'
        ))

        fig.update_layout(
            xaxis_title='Координата по Х (UTM)',
            yaxis_title='Координата по Y (UTM)',
            showlegend=False,
            width=900,
            height=800,
            margin=dict(t=30),
            title_text=f"Расположение скважин",
            title_x=0.5,
            title_y=0.98,
        )

        config = {
            'displaylogo': False,
            'displayModeBar': False
        }
        # Создание HTML-контента для графика
        plotly_js_cdn = "https://cdn.plot.ly/plotly-latest.min.js"
        raw_html = f'<script src="{plotly_js_cdn}"></script>' + fig.to_html(full_html=False, include_plotlyjs=False, config = config)

        # Установка HTML-контента в QWebEngineView
        self.ui.webEngineView.setHtml(raw_html)


    def plot_full_graph(self):
        if not self.file_path:
            QMessageBox.warning(self, "Файл не выбран", "Пожалуйста, выберите файл LAS перед построением графика.")
            return

        try:
            well_name = os.path.basename(self.file_path).replace(".las", "")
            selected_columns = self.get_selected_columns()
            print("Загрузка данных скважины...")
            force_well = Well(filename=self.file_path, path=None)
            test_well = pd.DataFrame(force_well.df())
            print(test_well.head(5))
            print("Данные загружены успешно.")
            test_well = test_well.dropna(subset=['FORCE_2020_LITHOFACIES_LITHOLOGY'])
            max_rows = 10000
            if len(test_well) > max_rows:
                print(f"Сокращение данных с {len(test_well)} строк до {max_rows} строк...")
                test_well = test_well.tail(max_rows)
                print(f"Сокращенные данные до {len(test_well)} строк.")

            if not selected_columns:
                QMessageBox.information(self, "Колонки не выбраны", "Пожалуйста, выберите хотя бы одну колонку для построения графика.")
                return

            print("Генерация полного графика...")
            fig = cwp.plot_logs(
                df=test_well,
                logs=selected_columns,
                show_fig=False
            )
            fig.update_layout(
                width=820,
                height=720,
                title_text=f"Скважина {well_name}",
                title_x=0.5,
                title_y=0.98,
                margin=dict(t=55)
            )
            print("График сгенерирован.")

            print("Преобразование графика в HTML...")
            config = {
                'displaylogo': False,
                'displayModeBar': False
            }
            plotly_js_path = "./plotly-2.32.0.min.js"

            raw_html = plot(fig, include_plotlyjs=False, output_type='div', config=config)
            raw_html = f'<script src="{plotly_js_path}"></script>' + raw_html

            with open("log_plot.html", "w") as debug_file:
                debug_file.write(raw_html)

            print("HTML сгенерирован.")

            print("Загрузка HTML в QWebEngineView...")
            self.ui.webEngineView.setHtml(raw_html, QUrl.fromLocalFile(os.path.abspath("log_plot.html")))

        except Exception as e:
            print("Ошибка при построении графика:", str(e))

    def get_lithology_predictions2(self):
        if not self.file_path:
            QMessageBox.warning(self, "Файл не выбран", "Пожалуйста, выберите файл LAS перед построением графика.")
            return

        try:
            print("Загрузка данных скважины...")
            force_well = Well(filename=self.file_path, path=None)
            test_well = pd.DataFrame(force_well.df())
            print(test_well.columns)
            print("Данные загружены успешно.")
            self.predictions = self.lithology_model.predict(test_well[Constants.FEATURES])
            test_well = test_well.iloc[5000:]
            test_well = test_well.dropna(subset=['FORCE_2020_LITHOFACIES_LITHOLOGY'])
            max_rows = 10000
            if len(test_well) > max_rows:
                print(f"Сокращение данных с {len(test_well)} строк до {max_rows} строк...")
                test_well = test_well.tail(max_rows)
                print(f"Сокращенные данные до {len(test_well)} строк.")

            print("Получение предсказаний...")
            selected_columns = self.get_selected_columns()
            if not set(Constants.FEATURES).issubset(test_well.columns):
                missing = list(set(Constants.FEATURES) - set(test_well.columns))
                QMessageBox.critical(self, "Недостающие колонки",
                                     f"Отсутствуют следующие необходимые столбцы: {', '.join(missing)}")

            predictions = self.lithology_model.predict(test_well[Constants.FEATURES])

            if predictions is not None:
                test_well['Predicted_Class'] = predictions

                test_well['Facies'] = test_well['Predicted_Class'].map(Constants.LITHOLOGY_KEYS).map(
                    Constants.LITHOLOGY_NUMBERS)
                test_well['Predicted Facies'] = test_well['FORCE_2020_LITHOFACIES_LITHOLOGY'].map(
                    Constants.LITHOLOGY_NUMBERS)
                print("Предсказания обработаны")
                self.test_well = test_well
                unique_facies = test_well['Predicted Facies'].unique()
                unique_lithology_description = {key: self.lithology_description.dictionary[key] for key in unique_facies
                                                if
                                                key in self.lithology_description.dictionary}
                print("Генерация графика...")
                print(unique_lithology_description)
                fig = cwp.plot_logs(
                    df=test_well,
                    logs=selected_columns,
                    lithology_logs=['Facies'],
                    lithology_description=LithologyDescription(unique_lithology_description),
                    show_fig=False
                )
                well_name = os.path.basename(self.file_path).replace(".las", "")
                fig_with_legend = add_legend_to_figure(fig, self.lithology_description, well_name)
                config = {
                    'displaylogo': False,
                    'displayModeBar': False
                }
                print("Преобразование графика в HTML...")
                plotly_js_path = "./plotly-2.32.0.min.js"

                raw_html = plot(fig_with_legend, include_plotlyjs=False, output_type='div', config=config)
                raw_html = f'<script src="{plotly_js_path}"></script>' + raw_html

                with open("prediction.html", "w") as debug_file:
                    debug_file.write(raw_html)

                print("HTML сгенерирован.")

                print("Загрузка HTML в QWebEngineView...")
                self.ui.webEngineView.setHtml(raw_html, QUrl.fromLocalFile(os.path.abspath("prediction.html")))

        except Exception as e:
            print("Ошибка при выполнении предсказания:", str(e))
    def get_lithology_predictions(self):
        if not self.file_path:
            QMessageBox.warning(self, "Файл не выбран", "Пожалуйста, выберите файл LAS перед построением графика.")
            return

        try:
            url = 'http://127.0.0.1:8081/predict'
            files = {'file': open(self.file_path, 'rb')}
            response = requests.post(url, files=files)

            if response.status_code == 200:
                # Сохранение файла локально
                base_name = os.path.basename(self.file_path)
                name, ext = os.path.splitext(base_name)
                predicted_file_path = f'{name}_predicted{ext}'
                with open(predicted_file_path, 'wb') as f:
                    f.write(response.content)

                print(f"Предсказанный файл сохранен как '{predicted_file_path}'")

                # Создание объекта Well из предсказанного файла
                force_well = Well(filename=predicted_file_path, path=None)
                test_well = pd.DataFrame(force_well.df())
                print(test_well.columns)
                print("Данные загружены успешно.")

                # Проверка наличия предсказаний
                if 'PREDICTIONS' in test_well.columns:
                    print("Предсказания успешно сохранены в базу данных.")

                    # Обработка данных
                    self.predictions = test_well['PREDICTIONS']
                    test_well = test_well.iloc[5000:]
                    test_well = test_well.dropna(subset=['FORCE_2020_LITHOFACIES_LITHOLOGY'])
                    max_rows = 10000
                    if len(test_well) > max_rows:
                        print(f"Сокращение данных с {len(test_well)} строк до {max_rows} строк...")
                        test_well = test_well.tail(max_rows)
                        print(f"Сокращенные данные до {len(test_well)} строк.")

                    print("Получение предсказаний...")
                    selected_columns = self.get_selected_columns()

                    # Обработка и визуализация предсказаний
                    test_well['Facies'] = test_well['PREDICTIONS'].map(Constants.LITHOLOGY_KEYS).map(
                        Constants.LITHOLOGY_NUMBERS)
                    test_well['Predicted Facies'] = test_well['FORCE_2020_LITHOFACIES_LITHOLOGY'].map(
                        Constants.LITHOLOGY_NUMBERS)
                    print(test_well['Facies'])
                    print("Предсказания обработаны")
                    self.test_well = test_well
                    unique_facies = test_well['Predicted Facies'].unique()
                    unique_lithology_description = {key: self.lithology_description.dictionary[key] for key in
                                                    unique_facies if
                                                    key in self.lithology_description.dictionary}
                    print("Генерация графика...")
                    print(unique_lithology_description)
                    fig = cwp.plot_logs(
                        df=test_well,
                        logs=selected_columns,
                        lithology_logs=['Facies'],
                        lithology_description=LithologyDescription(unique_lithology_description),
                        show_fig=False
                    )
                    well_name = os.path.basename(self.file_path).replace(".las", "")
                    fig_with_legend = add_legend_to_figure(fig, self.lithology_description, well_name)
                    config = {
                        'displaylogo': False,
                        'displayModeBar': False
                    }
                    print("Преобразование графика в HTML...")
                    plotly_js_path = "./plotly-2.32.0.min.js"

                    raw_html = plot(fig_with_legend, include_plotlyjs=False, output_type='div', config=config)
                    raw_html = f'<script src="{plotly_js_path}"></script>' + raw_html

                    with open("prediction.html", "w") as debug_file:
                        debug_file.write(raw_html)

                    print("HTML сгенерирован.")

                    print("Загрузка HTML в QWebEngineView...")
                    self.ui.webEngineView.setHtml(raw_html, QUrl.fromLocalFile(os.path.abspath("prediction.html")))

            else:
                QMessageBox.critical(self, "Ошибка предсказания",
                                     f"Ошибка при выполнении предсказания: {response.status_code}")

        except Exception as e:
            print(f"Ошибка при выполнении предсказания: {str(e)}")
            QMessageBox.critical(self, "Ошибка предсказания", f"Ошибка при выполнении предсказания: {str(e)}")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())